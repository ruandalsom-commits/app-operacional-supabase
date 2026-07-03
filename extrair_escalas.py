import urllib.request
import urllib.parse
import json
import http.cookiejar
import os
import time
from datetime import date
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACOES — CONTAS E REGIOES
# ============================================================

CONTAS = [
    {
        "email": "ruan.dalsom@masterdeliveryexpress.com.br",
        "senha": "Ruankz100%",
        "regioes": [
            {"nome": "Sao Paulo", "uuid": "3841c245-fac1-40a6-8b8f-8d6876447a6d"},
        ]
    },
    {
        "email": "ruan.dalsom@crmasterfilial1rj.com.br",
        "senha": "Ruankz100%",
        "regioes": [
            {"nome": "Rio - Madureira",    "uuid": "2800ed66-03d2-4877-880a-8de7ad2051cd"},
            {"nome": "Rio - Barra",        "uuid": "30a4c365-0ff0-4db7-8691-82e8470b3820"},
            {"nome": "Rio - Zona Sul",     "uuid": "4506e189-2547-4b77-a711-304f519da4d0"},
            {"nome": "Niteroi",            "uuid": "2721a164-0a92-4106-8aca-d68b6e4de9b3"},
            {"nome": "Rio - Campo Grande", "uuid": "82c43d58-08d5-4e15-9656-d4cf41b67855"},
        ]
    },
    {
        "email": "ruan.dalson@entregoaldeota.com.br",
        "senha": "Ruankz100%",
        "regioes": [
            {"nome": "Fortaleza", "uuid": "5ed2c46b-f439-45f6-a6c1-a4ff03519fdd"},
        ]
    },
]

ARQUIVO_SAIDA     = "escalas.xlsx"
ARQUIVO_JWT_BASE  = "jwt_escalas_{}.txt"
INTERVALO_MINUTOS = 10

# ============================================================
# URLS
# ============================================================

URL_VALIDATE = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/validate"
URL_TOKEN    = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/token"
URL_REFRESH  = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/token/refresh"
URL_ESCALAS  = "https://api.entregolog.com/logistics-web-bff/operation/v2.0/logistics-operator/shift-schedules"

LIMITE = 50

HEADERS_BASE = {
    "accept":                 "application/json, text/plain, */*",
    "accept-language":        "pt",
    "origin":                 "https://franqueado.entregolog.com",
    "referer":                "https://franqueado.entregolog.com/",
    "user-agent":             "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    "x-cookie-login":         "true",
    "x-country":              "BR",
    "x-ifood-logistics-auth": "true",
    "x-timezone":             "America/Sao_Paulo",
}

COLUNAS = ["Regiao", "Data", "Turno", "Praca", "Subpraca", "Logados", "Slots", "% Logados", "Horario Coleta", "Modal"]

# ============================================================
# SESSAO POR CONTA
# ============================================================

class Sessao:
    def __init__(self, conta):
        self.email      = conta["email"]
        self.senha      = conta["senha"]
        self.regioes    = conta["regioes"]
        self.jwt_file = ARQUIVO_JWT_BASE.format(self.email.replace("@", "_").replace(".", "_"))
        self.cookie_jar = http.cookiejar.CookieJar()
        self.opener     = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar)
        )

    def fazer_request(self, url, headers_extra=None, body=None):
        h = {**HEADERS_BASE, **(headers_extra or {})}
        req = urllib.request.Request(url, headers=h, data=body)
        with self.opener.open(req) as resp:
            raw = resp.read()
            try:
                return json.loads(raw.decode("utf-8"))
            except Exception:
                return {}

    def get_cookie(self, nome):
        for cookie in self.cookie_jar:
            if cookie.name == nome:
                return cookie.value
        return None

    def carregar_jwt(self):
        if os.path.exists(self.jwt_file):
            with open(self.jwt_file, "r") as f:
                jwt = f.read().strip()
                if jwt:
                    return jwt
        return None

    def salvar_jwt(self, jwt):
        with open(self.jwt_file, "w") as f:
            f.write(jwt)

    def solicitar_codigo(self):
        body = json.dumps({"email": self.email, "password": self.senha}).encode("utf-8")
        self.fazer_request(
            URL_VALIDATE,
            headers_extra={
                "content-type":    "application/json; charset=UTF-8",
                "x-captcha-token": "",
            },
            body=body
        )

    def autenticar(self, codigo, jwt_antigo):
        body = json.dumps({"email": self.email, "code": codigo}).encode("utf-8")
        h = {
            "content-type": "application/json; charset=UTF-8",
            "cookie":       f"entregolog_jwt={jwt_antigo}",
        }
        req = urllib.request.Request(URL_TOKEN, headers={**HEADERS_BASE, **h}, data=body)
        with self.opener.open(req) as resp:
            resp.read()
        jwt_novo = self.get_cookie("entregolog_jwt")
        if jwt_novo:
            self.salvar_jwt(jwt_novo)
            return jwt_novo
        return None

    def renovar_jwt(self):
        refresh = self.get_cookie("entregolog_refresh_jwt")
        if not refresh:
            return False
        req = urllib.request.Request(
            URL_REFRESH,
            headers={**HEADERS_BASE, "cookie": f"entregolog_refresh_jwt={refresh}"},
            data=b""
        )
        req.get_method = lambda: "POST"
        with self.opener.open(req) as resp:
            resp.read()
        jwt_novo = self.get_cookie("entregolog_jwt")
        if jwt_novo:
            self.salvar_jwt(jwt_novo)
            return True
        return False

    def buscar_pagina(self, page, data_inicio, data_fim):
        params = urllib.parse.urlencode({
            "dateFrom": data_inicio,
            "dateTo":   data_fim,
            "page":     page,
            "limit":    LIMITE,
        })
        return self.fazer_request(f"{URL_ESCALAS}?{params}")

    def extrair_todas(self, data_inicio, data_fim):
        dados     = self.buscar_pagina(0, data_inicio, data_fim)
        total     = dados["total"]
        registros = dados["values"]
        paginas   = (total + LIMITE - 1) // LIMITE
        for p in range(1, paginas):
            registros.extend(self.buscar_pagina(p, data_inicio, data_fim)["values"])
        return registros

# ============================================================
# LOGICA DE PRACA E SUBPRACA
# ============================================================

def calcular_praca_subpraca(r):
    regiao = r["region"]["name"] or ""
    sub    = r["subRegion"]["name"] or ""
    origem = r["origin"]["name"] or ""
    if origem:
        return "SP DEDICADO", origem
    return regiao.upper(), sub if sub else "LIVRE"

# ============================================================
# XLSX
# ============================================================

def criar_cabecalho(ws):
    ws.append(COLUNAS)
    header_fill = PatternFill("solid", fgColor="2F75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(COLUNAS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

def ajustar_colunas(ws):
    larguras = [18, 12, 28, 16, 30, 10, 8, 12, 18, 14]
    for i, largura in enumerate(larguras, 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

def carregar_chaves_existentes(ws):
    chaves = set()
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            # A chave agora inclui o Modal (row[9])
            modal_val = str(row[9]) if len(row) > 9 and row[9] else "Todos"
            chave = (str(row[0]), str(row[1]), str(row[2]), str(row[3]), str(row[4]), str(row[8]), modal_val)
            chaves.add(chave)
    return chaves

def salvar_xlsx(registros, horario_coleta, nome_regiao):
    if os.path.exists(ARQUIVO_SAIDA):
        wb = load_workbook(ARQUIVO_SAIDA)
        ws = wb.active
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = "Escalas"
        criar_cabecalho(ws)
        ajustar_colunas(ws)

    chaves_existentes = carregar_chaves_existentes(ws)
    novas = 0
    fill_par   = PatternFill("solid", fgColor="DEEAF1")
    fill_impar = PatternFill("solid", fgColor="FFFFFF")

    for r in registros:
        praca, subpraca = calcular_praca_subpraca(r)
        slots   = r["maxRegularDrivers"]
        logados = r["reservedRegularDrivers"]
        pct     = f"{round(logados / slots * 100, 1)}%" if slots > 0 else "0%"
        
        # Extrai o modal da lista 'modals' vinda do JSON
        modals_list = r.get("modals", [])
        if len(modals_list) == 1:
            modal = modals_list[0].get("name", "Todos")
        elif len(modals_list) > 1:
            modal = "Todos"
        else:
            modal = "Todos"
        
        chave   = (nome_regiao, r["date"], r["shift"]["name"], praca, subpraca, horario_coleta, modal)

        if chave in chaves_existentes:
            continue

        linha = ws.max_row + 1
        fill  = fill_par if linha % 2 == 0 else fill_impar

        valores = [
            nome_regiao,
            r["date"],
            r["shift"]["name"],
            praca,
            subpraca,
            logados,
            slots,
            pct,
            horario_coleta,
            modal,
        ]

        for col, valor in enumerate(valores, 1):
            cell = ws.cell(row=linha, column=col, value=valor)
            cell.fill = fill
            cell.alignment = Alignment(horizontal="center")

        chaves_existentes.add(chave)
        novas += 1

    wb.save(ARQUIVO_SAIDA)
    return novas

# ============================================================
# COLETA
# ============================================================

def coletar(sessoes):
    hoje           = date.today().strftime("%Y-%m-%d")
    horario_coleta = time.strftime("%Y-%m-%d %H:%M")
    print(f"[{horario_coleta}] Coletando escalas...")

    for sessao in sessoes:
        for regiao in sessao.regioes:
            try:
                registros = sessao.extrair_todas(hoje, hoje)
                novas     = salvar_xlsx(registros, horario_coleta, regiao["nome"])
                print(f"  {regiao['nome']}: {len(registros)} registros, {novas} novos.")
            except Exception as e:
                if "401" in str(e):
                    print(f"  {regiao['nome']}: JWT expirado, renovando...")
                    try:
                        sessao.renovar_jwt()
                        registros = sessao.extrair_todas(hoje, hoje)
                        novas     = salvar_xlsx(registros, horario_coleta, regiao["nome"])
                        print(f"  {regiao['nome']}: {len(registros)} registros, {novas} novos.")
                    except Exception as e2:
                        print(f"  {regiao['nome']}: Falha: {e2}")
                else:
                    print(f"  {regiao['nome']}: Erro: {e}")

# ============================================================
# EXECUCAO
# ============================================================

if __name__ == "__main__":
    try:
        sessoes = [Sessao(c) for c in CONTAS]

        for sessao in sessoes:
            jwt_antigo = sessao.carregar_jwt()
            if not jwt_antigo:
                raise Exception(f"JWT nao encontrado para {sessao.email}.")

            print(f"\nConta: {sessao.email}")
            print("Solicitando codigo...")
            sessao.solicitar_codigo()
            print("Verifique o email e digite o codigo.")
            print("Voce tem 2 minutos!")

            codigo = input("Digite o codigo: ").strip()
            jwt_novo = sessao.autenticar(codigo, jwt_antigo)
            if not jwt_novo:
                raise Exception(f"Autenticacao falhou para {sessao.email}. Rode novamente.")
            print(f"Login OK: {sessao.email}\n")

            if sessao != sessoes[-1]:
                print("Aguarde 5 segundos antes da proxima conta...")
                time.sleep(5)

        print(f"Iniciando coleta automatica a cada {INTERVALO_MINUTOS} minutos.")
        print("Deixe essa janela aberta. Para parar feche a janela ou Ctrl+C.\n")

        while True:
            coletar(sessoes)
            print(f"Proxima coleta em {INTERVALO_MINUTOS} minutos...\n")
            time.sleep(INTERVALO_MINUTOS * 60)

    except KeyboardInterrupt:
        print("\nColeta encerrada.")
    except Exception as e:
        print(f"\nErro: {e}")
        input("\nPressione Enter para fechar...")