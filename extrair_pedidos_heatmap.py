import urllib.request
import urllib.parse
import urllib.error
import json
import http.cookiejar
import os
import time
from datetime import datetime, timedelta
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

# ============================================================
# CONFIGURACAO — CONTA DE TESTE (SAO PAULO)
# ============================================================

EMAIL  = "ruan.dalsom@masterdeliveryexpress.com.br"
SENHA  = "Ruankz100%"
REGIAO = {"nome": "Sao Paulo", "uuid": "3841c245-fac1-40a6-8b8f-8d6876447a6d"}

JWT_FILE     = "jwt_pedidos_sp.txt"
ARQUIVO_XLSX = "pedidos_heatmap.xlsx"

# ============================================================
# URLS
# ============================================================

URL_VALIDATE = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/validate"
URL_TOKEN    = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/token"
URL_REFRESH  = "https://api.entregolog.com/logistics-web-bff/operation/users/authentication/token/refresh"
URL_ORDERS   = "https://api.entregolog.com/logistics-web-bff/operation/logistics-operator/orders"
URL_DETALHE_PEDIDO = "https://api.entregolog.com/logistics-web-bff/operation/logistics-operator/orders/{order_id}"

LIMITE = 50

# Pausa entre requests de detalhe, pra nao sobrecarregar / evitar bloqueio anti-bot
PAUSA_ENTRE_DETALHES = 0.35

HEADERS_BASE = {
    "accept":                 "application/json, text/plain, */*",
    "accept-language":        "pt",
    "origin":                 "https://franqueado.entregolog.com",
    "referer":                "https://franqueado.entregolog.com/",
    "user-agent":             "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 (KHTML, like Gecko) Chrome/146.0.0.0 Safari/537.36",
    "x-cookie-login":         "true",
    "x-country":              "BR",
    "x-ifood-logistics-auth": "true",
    "x-timezone":             "America/Sao_Paulo",
}

# ============================================================
# SESSAO / AUTENTICACAO
# ============================================================

class Sessao:
    def __init__(self, email, senha):
        self.email      = email
        self.senha      = senha
        self.cookie_jar = http.cookiejar.CookieJar()
        self.opener     = urllib.request.build_opener(
            urllib.request.HTTPCookieProcessor(self.cookie_jar)
        )

    def fazer_request(self, url, headers_extra=None, body=None, method=None):
        h = {**HEADERS_BASE, **(headers_extra or {})}
        req = urllib.request.Request(url, headers=h, data=body)
        if method:
            req.get_method = lambda: method
        with self.opener.open(req) as resp:
            raw = resp.read()
            try:
                return json.loads(raw.decode("utf-8"))
            except Exception:
                return {}

    def fazer_request_com_renovacao(self, url, headers_extra=None, body=None, method=None):
        """
        Faz a request normalmente. Se der 401 (token expirado), renova o JWT
        usando o refresh token e tenta de novo, uma unica vez.
        """
        try:
            return self.fazer_request(url, headers_extra=headers_extra, body=body, method=method)
        except urllib.error.HTTPError as e:
            if e.code == 401:
                renovado = self.renovar_jwt()
                if renovado:
                    return self.fazer_request(url, headers_extra=headers_extra, body=body, method=method)
            raise

    def get_cookie(self, nome):
        for cookie in self.cookie_jar:
            if cookie.name == nome:
                return cookie.value
        return None

    def carregar_jwt(self):
        if os.path.exists(JWT_FILE):
            with open(JWT_FILE, "r") as f:
                jwt = f.read().strip()
                if jwt:
                    return jwt
        return None

    def salvar_jwt(self, jwt):
        with open(JWT_FILE, "w") as f:
            f.write(jwt)

    def solicitar_codigo(self):
        body = json.dumps({"email": self.email, "password": self.senha}).encode("utf-8")
        self.fazer_request(
            URL_VALIDATE,
            headers_extra={
                "content-type":    "application/json; charset=UTF-8",
                "x-captcha-token": "",
            },
            body=body
        )

    def autenticar(self, codigo, jwt_antigo=""):
        body = json.dumps({"email": self.email, "code": codigo}).encode("utf-8")
        h = {"content-type": "application/json; charset=UTF-8"}
        if jwt_antigo:
            h["cookie"] = f"entregolog_jwt={jwt_antigo}"
        req = urllib.request.Request(URL_TOKEN, headers={**HEADERS_BASE, **h}, data=body)
        with self.opener.open(req) as resp:
            resp.read()
        jwt_novo = self.get_cookie("entregolog_jwt")
        if jwt_novo:
            self.salvar_jwt(jwt_novo)
            return jwt_novo
        return None

    def renovar_jwt(self):
        refresh = self.get_cookie("entregolog_refresh_jwt")
        if not refresh:
            return False
        req = urllib.request.Request(
            URL_REFRESH,
            headers={**HEADERS_BASE, "cookie": f"entregolog_refresh_jwt={refresh}"},
            data=b""
        )
        req.get_method = lambda: "POST"
        with self.opener.open(req) as resp:
            resp.read()
        jwt_novo = self.get_cookie("entregolog_jwt")
        if jwt_novo:
            self.salvar_jwt(jwt_novo)
            return True
        return False

    # --------------------------------------------------------
    # LISTAGEM DE PEDIDOS (TODAS AS PAGINAS)
    # --------------------------------------------------------

    def buscar_pagina_pedidos(self, page, data_inicio_utc, data_fim_utc):
        params = urllib.parse.urlencode({
            "page":            page,
            "size":            LIMITE,
            "states":          "COMPLETED",
            "sort":            "deliveryTimeWindowEnd,desc",
            "createdDateFrom": data_inicio_utc,
            "createdDateTo":   data_fim_utc,
        })
        params += "&states=CANCELLED&states=RETURNED"
        return self.fazer_request_com_renovacao(f"{URL_ORDERS}?{params}")

    def extrair_todos_pedidos(self, data_inicio_utc, data_fim_utc):
        dados     = self.buscar_pagina_pedidos(0, data_inicio_utc, data_fim_utc)
        total     = dados.get("total", 0)
        registros = dados.get("content", [])
        paginas   = (total + LIMITE - 1) // LIMITE
        print(f"Total de pedidos no periodo: {total} ({paginas} paginas)")

        for p in range(1, paginas):
            print(f"  Buscando pagina {p + 1}/{paginas}...")
            pagina = self.buscar_pagina_pedidos(p, data_inicio_utc, data_fim_utc)
            registros.extend(pagina.get("content", []))

        return registros

    # --------------------------------------------------------
    # DETALHE DO PEDIDO (lat/lng)
    # --------------------------------------------------------

    def buscar_detalhe(self, order_id):
        url = URL_DETALHE_PEDIDO.format(order_id=order_id)
        return self.fazer_request_com_renovacao(url)


# ============================================================
# EXTRACAO DOS CAMPOS RELEVANTES
# ============================================================

def extrair_campos_heatmap(detalhe):
    origin      = detalhe.get("origin", {}) or {}
    origin_addr = origin.get("address", {}) or {}
    dest        = detalhe.get("destination", {}) or {}
    dest_addr   = dest.get("address", {}) or {}

    return {
        "order_id":        detalhe.get("orderId") or detalhe.get("id"),
        "delivery_code":   detalhe.get("deliveryCode"),
        "order_reference": detalhe.get("orderReference"),
        "region_name":     (detalhe.get("region") or {}).get("name"),
        "sub_region_name": (detalhe.get("subRegion") or {}).get("name"),
        "cluster_name":    (detalhe.get("cluster") or {}).get("name"),
        "state":           (detalhe.get("state") or {}).get("current"),
        "created_at":      detalhe.get("createdAt"),
        "origin_name":     origin.get("name"),
        "origin_lat":      origin_addr.get("latitude"),
        "origin_lng":      origin_addr.get("longitude"),
        "destination_lat": dest_addr.get("latitude"),
        "destination_lng": dest_addr.get("longitude"),
    }


# ============================================================
# XLSX
# ============================================================

COLUNAS_DADOS = [
    "order_id", "delivery_code", "order_reference", "region_name",
    "sub_region_name", "cluster_name", "state", "created_at",
    "origin_name", "origin_lat", "origin_lng", "destination_lat", "destination_lng",
]

COLUNAS_ERROS = ["delivery_code", "order_id", "erro"]


def montar_planilha(resultados, erros):
    wb = Workbook()

    # --- Aba de dados ---
    ws = wb.active
    ws.title = "Pedidos"
    ws.append(COLUNAS_DADOS)

    header_fill = PatternFill("solid", fgColor="2F75B6")
    header_font = Font(bold=True, color="FFFFFF")
    for col in range(1, len(COLUNAS_DADOS) + 1):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in resultados:
        ws.append([item.get(c) for c in COLUNAS_DADOS])

    for i, largura in enumerate([38, 14, 38, 14, 22, 30, 16, 26, 22, 14, 14, 14, 14], 1):
        ws.column_dimensions[get_column_letter(i)].width = largura

    # --- Aba de erros ---
    ws_err = wb.create_sheet("Erros")
    ws_err.append(COLUNAS_ERROS)
    for col in range(1, len(COLUNAS_ERROS) + 1):
        cell = ws_err.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for item in erros:
        ws_err.append([item.get(c) for c in COLUNAS_ERROS])

    for i, largura in enumerate([14, 38, 60], 1):
        ws_err.column_dimensions[get_column_letter(i)].width = largura

    wb.save(ARQUIVO_XLSX)


# ============================================================
# EXECUCAO
# ============================================================

if __name__ == "__main__":
    sessao = Sessao(EMAIL, SENHA)

    jwt_antigo = sessao.carregar_jwt()
    if not jwt_antigo:
        print("Nenhum JWT salvo ainda. Sera necessario fazer login com codigo por e-mail.")

    print(f"\nConta: {EMAIL}")
    print("Solicitando codigo de login...")
    sessao.solicitar_codigo()
    print("Verifique o email e digite o codigo. Voce tem 2 minutos!")

    codigo = input("Digite o codigo: ").strip()
    jwt_novo = sessao.autenticar(codigo, jwt_antigo or "")
    if not jwt_novo:
        raise SystemExit("Autenticacao falhou. Rode novamente.")
    print("Login OK.\n")

    # --------------------------------------------------------
    # Range: dia de ontem (BRT -> UTC)
    # --------------------------------------------------------
    hoje_brt         = datetime.now().replace(hour=0, minute=0, second=0, microsecond=0)
    ontem_inicio_brt = hoje_brt - timedelta(days=1)
    ontem_fim_brt    = hoje_brt - timedelta(minutes=5)

    ontem_inicio_utc = (ontem_inicio_brt + timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M:%S.000Z")
    ontem_fim_utc    = (ontem_fim_brt + timedelta(hours=3)).strftime("%Y-%m-%dT%H:%M:%S.000Z")

    print(f"Buscando pedidos de {ontem_inicio_utc} ate {ontem_fim_utc} (UTC)...\n")

    registros = sessao.extrair_todos_pedidos(ontem_inicio_utc, ontem_fim_utc)
    print(f"\n{len(registros)} pedidos encontrados. Buscando detalhes...\n")

    resultados = []
    erros = []

    for i, r in enumerate(registros, 1):
        order_id = r["orderId"]
        try:
            detalhe = sessao.buscar_detalhe(order_id)
            campos = extrair_campos_heatmap(detalhe)
            resultados.append(campos)
        except urllib.error.HTTPError as e:
            descricao = f"HTTP {e.code}" + (" (token expirado, renovacao falhou)" if e.code == 401 else "")
            erros.append({
                "delivery_code": r.get("deliveryCode"),
                "order_id": order_id,
                "erro": descricao,
            })
        except Exception as e:
            erros.append({
                "delivery_code": r.get("deliveryCode"),
                "order_id": order_id,
                "erro": str(e),
            })

        if i % 25 == 0 or i == len(registros):
            print(f"  Processados {i}/{len(registros)} (ok: {len(resultados)}, erros: {len(erros)})")

        time.sleep(PAUSA_ENTRE_DETALHES)

    print(f"\nFinalizado: {len(resultados)} pedidos com sucesso, {len(erros)} com erro.")

    montar_planilha(resultados, erros)
    print(f"Planilha salva em: {os.path.abspath(ARQUIVO_XLSX)}")